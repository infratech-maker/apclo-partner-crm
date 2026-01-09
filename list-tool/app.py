"""Flaskアプリケーションファクトリ"""
from flask import Flask, send_from_directory, jsonify, request
from extensions import db, migrate
from config import config as config_dict
import os
import io


def create_app(config_name='default'):
    """アプリケーションファクトリ"""
    app = Flask(__name__)
    
    # 設定を読み込む
    if config_name not in config_dict:
        config_name = 'default'
    app.config.from_object(config_dict[config_name])
    
    # ngrokの警告ページをスキップするためのヘッダーを追加
    @app.after_request
    def after_request(response):
        # ngrokの警告ページをスキップするためのヘッダー
        response.headers['ngrok-skip-browser-warning'] = 'true'
        # CORSヘッダーを追加（必要に応じて）
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, ngrok-skip-browser-warning'
        return response
    
    # 拡張機能を初期化
    db.init_app(app)
    migrate.init_app(app, db)
    
    # データベーステーブルを自動作成（初回のみ）
    with app.app_context():
        try:
            db.create_all()
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"データベーステーブル作成をスキップ: {e}")
    
    # ルートを登録
    register_routes(app)
    
    # シェルコンテキストを設定
    @app.shell_context_processor
    def make_shell_context():
        from models import Store, User
        return {'db': db, 'Store': Store, 'User': User}
    
    # Celeryを初期化（Redisが利用可能な場合のみ）
    try:
        from tasks import make_celery
        celery_app = make_celery(app)
        
        # Celeryタスクを登録
        from tasks import register_tasks, register_enrichment_tasks
        register_tasks(celery_app)
        register_enrichment_tasks(celery_app)
    except Exception as e:
        # Redisが利用できない場合はCeleryを無効化
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f"Celeryの初期化をスキップしました: {e}")
    
    return app


def register_routes(app):
    """ルートを登録"""
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    
    # Webルート
    @app.route("/")
    def root():
        return send_from_directory(BASE_DIR, "list-tool.html")
    
    @app.route("/list-tool")
    def list_tool():
        return send_from_directory(BASE_DIR, "list-tool.html")
    
    @app.route("/login")
    def login():
        return send_from_directory(BASE_DIR, "login.html")
    
    @app.route("/dashboard")
    def dashboard():
        return send_from_directory(BASE_DIR, "dashboard.html")
    
    @app.route("/admin-dashboard")
    def admin_dashboard():
        return send_from_directory(BASE_DIR, "admin-dashboard.html")
    
    @app.route("/products")
    def products():
        return send_from_directory(BASE_DIR, "products.html")
    
    @app.route("/account-management")
    def account_management():
        return send_from_directory(BASE_DIR, "account-management.html")
    
    @app.route("/barius.html")
    def barius():
        return send_from_directory(BASE_DIR, "barius.html")
    
    @app.route("/barius")
    def barius_short():
        return send_from_directory(BASE_DIR, "barius.html")
    
    # 静的ファイル（画像）の配信
    @app.route("/<path:filename>")
    def serve_static(filename):
        """画像ファイルなどの静的ファイルを配信"""
        # 画像ファイルのみを配信（セキュリティ対策）
        if filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.svg', '.webp', '.JPG', '.JPEG', '.PNG', '.GIF', '.SVG', '.WEBP')):
            return send_from_directory(BASE_DIR, filename)
        else:
            from flask import abort
            abort(404)
    
    # APIルート
    @app.route("/api/health")
    def health():
        return jsonify({"status": "ok"})
    
    @app.route("/api/areas")
    def get_areas():
        """エリアリスト取得API"""
        try:
            AREAS = ["北海道", "東北", "関東", "中部", "近畿", "中国", "四国", "九州"]
            AREA_PREFECTURES = {
                "北海道": ["北海道"],
                "東北": ["青森", "岩手", "宮城", "秋田", "山形", "福島"],
                "関東": ["茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川"],
                "中部": ["新潟", "富山", "石川", "福井", "山梨", "長野", "岐阜", "静岡", "愛知"],
                "近畿": ["三重", "滋賀", "京都", "大阪", "兵庫", "奈良", "和歌山"],
                "中国": ["鳥取", "島根", "岡山", "広島", "山口"],
                "四国": ["徳島", "香川", "愛媛", "高知"],
                "九州": ["福岡", "佐賀", "長崎", "熊本", "大分", "宮崎", "鹿児島", "沖縄"]
            }
            return jsonify({
                "areas": AREAS,
                "area_prefectures": AREA_PREFECTURES
            })
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    @app.route("/api/prefectures")
    def get_prefectures():
        """都道府県リスト取得API"""
        try:
            PREFECTURES = [
                "北海道", "青森", "岩手", "宮城", "秋田", "山形", "福島",
                "茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川",
                "新潟", "富山", "石川", "福井", "山梨", "長野", "岐阜", "静岡", "愛知",
                "三重", "滋賀", "京都", "大阪", "兵庫", "奈良", "和歌山",
                "鳥取", "島根", "岡山", "広島", "山口",
                "徳島", "香川", "愛媛", "高知",
                "福岡", "佐賀", "長崎", "熊本", "大分", "宮崎", "鹿児島", "沖縄"
            ]
            return jsonify({"prefectures": PREFECTURES})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/cities")
    def get_cities():
        """指定した都道府県に属する市区町村リスト取得API"""
        try:
            from models import Store
            from sqlalchemy import func
            from sqlalchemy.exc import OperationalError

            prefecture = request.args.get("prefecture", "").strip()

            try:
                query = db.session.query(Store.city).filter(
                    Store.city.isnot(None),
                    Store.city != ""
                )

                # 住所の先頭に都道府県名が含まれている前提でフィルタ
                if prefecture:
                    query = query.filter(
                        Store.address.isnot(None),
                        Store.address != "",
                        Store.address.like(f"{prefecture}%"),
                    )

                cities_rows = query.distinct().all()
                cities = sorted({row[0] for row in cities_rows if row[0]})

                return jsonify({"cities": cities})
            except OperationalError as e:
                # テーブル未作成時は空リストを返す
                if "no such table" in str(e).lower():
                    return jsonify({"cities": []})
                raise
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    @app.route("/api/categories")
    def get_categories():
        """カテゴリリスト取得API（データベースから実際の値を取得し、純粋なカテゴリー名のみを抽出）"""
        try:
            from models import Store
            from sqlalchemy.exc import OperationalError
            import re
            
            def extract_category_names(category_value):
                """カテゴリー値から純粋なカテゴリー名を抽出
                例: 'あざみ野駅 100m / ドーナツ' -> ['ドーナツ']
                例: 'あびこ駅 313m / カフェ、スイーツ' -> ['カフェ', 'スイーツ']
                """
                if not category_value:
                    return []
                
                # 「/」で分割して、最後の部分（カテゴリー名部分）を取得
                if '/' in category_value:
                    parts = category_value.split('/')
                    category_part = parts[-1].strip()
                else:
                    category_part = category_value.strip()
                
                # カンマ区切りのカテゴリーを分割
                categories = [cat.strip() for cat in category_part.split('、') if cat.strip()]
                
                # さらにカンマでも分割（「カフェ, スイーツ」形式に対応）
                result = []
                for cat in categories:
                    result.extend([c.strip() for c in cat.split(',') if c.strip()])
                
                return result
            
            try:
                # データベースから実際のカテゴリー値を取得
                categories_query = db.session.query(Store.category).distinct().filter(
                    Store.category.isnot(None),
                    Store.category != ""
                ).order_by(Store.category)
                
                # すべてのカテゴリー値を取得
                all_category_values = [row[0] for row in categories_query.all() if row[0]]
                
                # 各カテゴリー値から純粋なカテゴリー名を抽出
                extracted_categories = set()
                for category_value in all_category_values:
                    category_names = extract_category_names(category_value)
                    extracted_categories.update(category_names)
                
                # ソートしてリストに変換
                categories_list = sorted(extracted_categories)
                
                return jsonify({
                    "categories": categories_list,
                    "category_groups": {}  # グループ化は不要なので空オブジェクト
                })
            except OperationalError as e:
                # テーブル未作成時は空リストを返す
                if "no such table" in str(e).lower():
                    return jsonify({
                        "categories": [],
                        "category_groups": {}
                    })
                raise
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    @app.route("/api/stats")
    def get_stats():
        """統計情報取得API"""
        try:
            from models import Store
            from sqlalchemy import func, and_, or_
            from sqlalchemy.exc import OperationalError
            
            # テーブルが存在するか確認
            latest_update_str = None
            try:
                # 全店舗数
                total_stores = db.session.query(func.count(func.distinct(Store.store_id))).scalar() or 0
                
                # 最終更新日時を取得（最も新しいupdated_at）
                latest_update = db.session.query(func.max(Store.updated_at)).scalar()
                if latest_update:
                    latest_update_str = latest_update.isoformat()
            except OperationalError as e:
                # エラー時もlatest_update_strはNoneのまま
                # テーブルが存在しない場合は0を返す
                if 'no such table' in str(e).lower():
                    return jsonify({
                        'total_stores': 0,
                        'total_with_opening': 0,
                        'with_opening_date_count': 0,
                        'remaining': 0,
                        'completed': 0,
                        'completion_rate': 0,
                        'with_phone': 0,
                        'with_website': 0,
                        'fully_completed': 0,
                        'fully_completed_with_opening': 0,
                        'latest_update': None,
                    })
                raise
                
            # 開店日ありの店舗数
            total_with_opening = db.session.query(
                func.count(func.distinct(Store.store_id))
            ).filter(Store.opening_date.isnot(None)).scalar() or 0

            # 補完が必要な件数
            remaining = db.session.query(func.count(func.distinct(Store.store_id))).filter(
                and_(
                    Store.opening_date.isnot(None),
                    Store.url.isnot(None),
                    Store.url != "",
                    or_(
                        Store.phone.is_(None),
                        Store.phone == "",
                        Store.closed_day.is_(None),
                        Store.closed_day == "",
                        Store.business_hours.is_(None),
                        Store.business_hours == "",
                        Store.transport.is_(None),
                        Store.transport == "",
                        Store.official_account.is_(None),
                        Store.official_account == "",
                    ),
                )
            ).scalar() or 0

            # 補完完了件数
            completed = total_with_opening - remaining if total_with_opening > 0 else 0
            completion_rate = (completed / total_with_opening * 100) if total_with_opening > 0 else 0

            # 電話番号あり
            with_phone = db.session.query(func.count(func.distinct(Store.store_id))).filter(
                Store.phone.isnot(None),
                Store.phone != "",
            ).scalar() or 0

            # ウェブサイトあり
            with_website = db.session.query(func.count(func.distinct(Store.store_id))).filter(
                Store.website.isnot(None),
                Store.website != "",
            ).scalar() or 0

            # 全項目完了
            fully_completed = db.session.query(func.count(func.distinct(Store.store_id))).filter(
                and_(
                    Store.opening_date.isnot(None),
                    Store.phone.isnot(None),
                    Store.phone != "",
                    Store.closed_day.isnot(None),
                    Store.closed_day != "",
                    Store.business_hours.isnot(None),
                    Store.business_hours != "",
                    Store.transport.isnot(None),
                    Store.transport != "",
                )
            ).scalar() or 0

            # 都市数（重複除く）
            cities_count = db.session.query(func.count(func.distinct(Store.city))).filter(
                Store.city.isnot(None),
                Store.city != "",
            ).scalar() or 0

            # 都道府県・エリア別統計用の定義（先に定義が必要）
            PREFECTURES = [
                "北海道",
                "青森",
                "岩手",
                "宮城",
                "秋田",
                "山形",
                "福島",
                "茨城",
                "栃木",
                "群馬",
                "埼玉",
                "千葉",
                "東京",
                "神奈川",
                "新潟",
                "富山",
                "石川",
                "福井",
                "山梨",
                "長野",
                "岐阜",
                "静岡",
                "愛知",
                "三重",
                "滋賀",
                "京都",
                "大阪",
                "兵庫",
                "奈良",
                "和歌山",
                "鳥取",
                "島根",
                "岡山",
                "広島",
                "山口",
                "徳島",
                "香川",
                "愛媛",
                "高知",
                "福岡",
                "佐賀",
                "長崎",
                "熊本",
                "大分",
                "宮崎",
                "鹿児島",
                "沖縄",
            ]
            
            # 市区町村別店舗数（上位20市区町村）
            # 住所から市区町村を抽出して集計
            import re
            
            def extract_city_from_address(address):
                """住所から市区町村を抽出"""
                if not address:
                    return None
                
                addr = address
                # 都道府県名を除去
                for pref in PREFECTURES:
                    if addr.startswith(pref):
                        addr = addr[len(pref):].lstrip('都府県')
                        break
                
                # 駅名と距離情報を除去（例: "池袋駅 396m"）
                addr = re.sub(r'[^都府県市区町村]*駅\s*\d+m\s*/?', '', addr)
                
                # カテゴリー情報を除去（例: "/ カテゴリー"）
                addr = re.sub(r'/\s*[^/]+$', '', addr)
                
                # 市区町村パターンを抽出
                # パターン1: "XX区", "XX市", "XX町", "XX村"（都道府県名の後）
                match = re.search(r'([^都府県市区町村]+[市区町村])', addr)
                if match:
                    city = match.group(1).strip()
                    # 余分な文字を除去
                    city = re.sub(r'^\s*[、,]\s*', '', city)
                    if city and len(city) > 1 and not city.startswith('駅'):
                        return city
                
                # パターン2: "XX郡XX町", "XX郡XX村"
                match = re.search(r'([^都府県市区町村]+郡[^市区町村]+[町村])', addr)
                if match:
                    city = match.group(1).strip()
                    if city and len(city) > 1:
                        return city
                
                return None
            
            # 住所から市区町村を抽出して集計
            address_rows = db.session.query(Store.address, Store.city).filter(
                Store.address.isnot(None),
                Store.address != "",
            ).all()
            
            city_stats = {}
            for addr, orig_city in address_rows:
                # まず住所から市区町村を抽出
                extracted_city = extract_city_from_address(addr)
                
                # 抽出できなかった場合は、元のcityフィールドを使用
                # ただし、都道府県名の場合は除外
                if not extracted_city:
                    if orig_city and orig_city not in PREFECTURES:
                        extracted_city = orig_city
                    else:
                        continue
                
                if extracted_city:
                    city_stats[extracted_city] = city_stats.get(extracted_city, 0) + 1
            
            # 店舗数の降順でソートして上位20を取得
            city_stats = dict(sorted(city_stats.items(), key=lambda x: x[1], reverse=True)[:20])

            # 都道府県・エリア別統計用の定義（既に上で定義済み）

            AREA_PREFECTURES = {
                "北海道": ["北海道"],
                "東北": ["青森", "岩手", "宮城", "秋田", "山形", "福島"],
                "関東": ["茨城", "栃木", "群馬", "埼玉", "千葉", "東京", "神奈川"],
                "中部": ["新潟", "富山", "石川", "福井", "山梨", "長野", "岐阜", "静岡", "愛知"],
                "近畿": ["三重", "滋賀", "京都", "大阪", "兵庫", "奈良", "和歌山"],
                "中国": ["鳥取", "島根", "岡山", "広島", "山口"],
                "四国": ["徳島", "香川", "愛媛", "高知"],
                "九州": ["福岡", "佐賀", "長崎", "熊本", "大分", "宮崎", "鹿児島", "沖縄"],
            }

            # 都道府県を判定して集計（住所と都市名の両方をチェック）
            prefecture_stats = {p: 0 for p in PREFECTURES}
            
            # 都市名から都道府県へのマッピング
            CITY_TO_PREFECTURE = {
                "東京": "東京",
                "神奈川": "神奈川",
                "千葉": "千葉",
                "埼玉": "埼玉",
                "大阪": "大阪",
                "神戸": "兵庫",
                "京都": "京都",
                "横浜": "神奈川",
                "川崎": "神奈川",
                "相模原": "神奈川",
                "さいたま": "埼玉",
                "川口": "埼玉",
                "船橋": "千葉",
                "市川": "千葉",
                "松山": "愛媛",
                "高知": "高知",
                "福島": "福島",
                "金沢": "石川",
                "宮崎": "宮崎",
                "鳥取": "鳥取",
            }
            
            # 店舗データを取得（住所と都市名の両方を使用）
            store_rows = db.session.query(Store.address, Store.city).filter(
                db.or_(
                    Store.address.isnot(None),
                    Store.city.isnot(None)
                ),
                db.or_(
                    Store.address != "",
                    Store.city != ""
                )
            ).all()
            
            for addr, city in store_rows:
                matched = False
                
                # まず都市名から都道府県を判定
                if city:
                    # 都市名が都道府県名と一致する場合
                    if city in PREFECTURES:
                        prefecture_stats[city] += 1
                        matched = True
                    # 都市名から都道府県をマッピング
                    elif city in CITY_TO_PREFECTURE:
                        pref = CITY_TO_PREFECTURE[city]
                        prefecture_stats[pref] += 1
                        matched = True
                
                # 都市名でマッチしなかった場合、住所から判定
                if not matched and addr:
                    for pref in PREFECTURES:
                        if addr.startswith(pref):
                            prefecture_stats[pref] += 1
                            break

            # エリア別に集計
            area_stats = {}
            for area, prefs in AREA_PREFECTURES.items():
                area_stats[area] = sum(prefecture_stats.get(p, 0) for p in prefs)

            # 取得率
            phone_rate = (with_phone / total_stores * 100) if total_stores > 0 else 0
            website_rate = (with_website / total_stores * 100) if total_stores > 0 else 0

            return jsonify(
                {
                    "total_stores": total_stores,
                    "total_with_opening": total_with_opening,
                    "with_opening_date_count": total_with_opening,
                    "remaining": remaining,
                    "completed": completed,
                    "completion_rate": completion_rate,
                    "with_phone": with_phone,
                    "with_website": with_website,
                    "fully_completed": fully_completed,
                    "fully_completed_with_opening": fully_completed,
                    "cities": cities_count,
                    "phone_rate": phone_rate,
                    "website_rate": website_rate,
                    "city_stats": city_stats,
                    "prefecture_stats": prefecture_stats,
                    "area_stats": area_stats,
                    "latest_update": latest_update_str,
                }
            )
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/stores")
    def get_stores():
        """店舗データ一覧取得API"""
        try:
            from models import Store
            from sqlalchemy import or_, and_, func
            from sqlalchemy.exc import OperationalError
            
            page = int(request.args.get("page", 1))
            per_page = int(request.args.get("per_page", 100))
            search = request.args.get("search", "").strip() or None
            
            try:
                query = db.session.query(Store)

                # クエリパラメータ
                search = request.args.get("search", "").strip() or None
                search_mode = request.args.get("search_mode", "AND")
                match_type = request.args.get("match_type", "partial")
                prefectures = request.args.getlist("prefectures")
                cities = request.args.getlist("cities")

                # キーワード検索（店舗名・住所・カテゴリ）
                if search:
                    keywords = [k for k in search.split() if k]
                    if keywords:
                        conditions = []
                        for kw in keywords:
                            pattern = f"%{kw}%" if match_type == "partial" else kw
                            conditions.append(
                                or_(
                                    Store.name.ilike(pattern),
                                    Store.address.ilike(pattern),
                                    Store.category.ilike(pattern),
                                )
                            )
                        if search_mode.upper() == "OR":
                            query = query.filter(or_(*conditions))
                        else:
                            for cond in conditions:
                                query = query.filter(cond)

                # 都道府県フィルター（住所の先頭に都道府県名が含まれている前提）
                if prefectures:
                    prefecture_filters = []
                    for pref in prefectures:
                        prefecture_filters.append(
                            and_(
                                Store.address.isnot(None),
                                Store.address != "",
                                Store.address.like(f"{pref}%"),
                            )
                        )
                    query = query.filter(or_(*prefecture_filters))

                # 市区町村フィルター
                if cities:
                    query = query.filter(
                        Store.city.isnot(None),
                        Store.city != "",
                        Store.city.in_(cities),
                    )

                total_count = query.count()
                stores = query.order_by(Store.store_id).offset((page - 1) * per_page).limit(per_page).all()
                
                return jsonify({
                    "stores": [store.to_dict() for store in stores],
                    "total": total_count,
                    "page": page,
                    "per_page": per_page,
                    "total_pages": (total_count + per_page - 1) // per_page if total_count > 0 else 0,
                })
            except OperationalError as e:
                # テーブルが存在しない場合は空のリストを返す
                if 'no such table' in str(e).lower():
                    return jsonify({
                        "stores": [],
                        "total": 0,
                        "page": page,
                        "per_page": per_page,
                        "total_pages": 0,
                    })
                raise
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/partner/saved-lists")
    def get_saved_lists():
        """保存済みリスト取得API（ダミー）"""
        return jsonify({"lists": []})
    
    @app.route("/api/login", methods=['POST'])
    def login_api():
        """ログインAPI"""
        try:
            from models import User
            from werkzeug.security import check_password_hash
            from sqlalchemy.exc import OperationalError
            from datetime import datetime
            
            data = request.get_json()
            
            if not data:
                return jsonify({"error": "リクエストボディが必要です"}), 400
            
            partner_code = data.get('partner_code', '').strip()
            password = data.get('password', '')
            
            if not partner_code or not password:
                return jsonify({"error": "パートナーIDとパスワードは必須です"}), 400
            
            try:
                # ユーザーを検索
                user = db.session.query(User).filter(
                    User.partner_code == partner_code
                ).first()
                
                if not user:
                    return jsonify({"error": "パートナーIDまたはパスワードが正しくありません"}), 401
                
                # ユーザーが無効な場合
                if not user.is_active:
                    return jsonify({"error": "このアカウントは無効です"}), 403
                
                # パスワードを検証
                if not check_password_hash(user.password_hash, password):
                    return jsonify({"error": "パートナーIDまたはパスワードが正しくありません"}), 401
                
                # 最終ログイン日時を更新
                user.last_login_at = datetime.utcnow()
                db.session.commit()
                
                # ログイン成功
                return jsonify({
                    "success": True,
                    "user": {
                        "user_id": str(user.user_id),
                        "partner_code": user.partner_code,
                        "name": user.name,
                        "user_type": user.user_type,
                        "is_admin": user.user_type == "admin",
                    }
                }), 200
                
            except OperationalError as e:
                if 'no such table' in str(e).lower():
                    # テーブルが存在しない場合は開発モードとして許可
                    # ただし、セキュリティ上、本番環境ではテーブルが必要
                    return jsonify({
                        "success": True,
                        "user": {
                            "partner_code": partner_code,
                            "name": "開発ユーザー",
                            "user_type": "partner",
                            "is_admin": partner_code.upper() == "ADMIN",
                        },
                        "warning": "データベーステーブルが存在しません。開発モードで動作しています。"
                    }), 200
                raise
                
        except Exception as e:
            import traceback
            db.session.rollback()
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/admin/users", methods=['GET', 'POST'])
    def admin_users():
        """ユーザー管理API（GET: 一覧取得、POST: 新規作成）"""
        try:
            from models import User
            from sqlalchemy.exc import OperationalError
            import uuid
            from werkzeug.security import generate_password_hash
            
            if request.method == 'GET':
                # クエリパラメータ
                user_type = request.args.get('user_type', '').strip() or None
                is_active = request.args.get('is_active', '').strip()
                
                try:
                    query = db.session.query(User)
                    
                    if user_type:
                        query = query.filter(User.user_type == user_type)
                    
                    if is_active:
                        is_active_bool = is_active.lower() == 'true'
                        query = query.filter(User.is_active == is_active_bool)
                    
                    users = query.order_by(User.created_at.desc()).all()
                    
                    return jsonify({
                        "users": [{
                            "user_id": str(user.user_id),
                            "partner_code": user.partner_code,
                            "name": user.name,
                            "email": user.email,
                            "phone": user.phone,
                            "organization": user.organization,
                            "user_type": user.user_type,
                            "is_agency": user.is_agency,
                            "is_active": user.is_active,
                            "created_at": user.created_at.isoformat() if user.created_at else None,
                            "updated_at": user.updated_at.isoformat() if user.updated_at else None,
                            "last_login_at": user.last_login_at.isoformat() if user.last_login_at else None,
                        } for user in users]
                    })
                except OperationalError as e:
                    if 'no such table' in str(e).lower():
                        return jsonify({"users": []})
                    raise
                    
            elif request.method == 'POST':
                # 新規ユーザー作成
                data = request.get_json()
                
                if not data:
                    return jsonify({"error": "リクエストボディが必要です"}), 400
                
                # 必須フィールドチェック
                required_fields = ['partner_code', 'password', 'name', 'user_type']
                for field in required_fields:
                    if field not in data or not data[field]:
                        return jsonify({"error": f"{field}は必須です"}), 400
                
                # パートナーコードの重複チェック
                existing = db.session.query(User).filter(
                    User.partner_code == data['partner_code']
                ).first()
                
                if existing:
                    return jsonify({"error": "このパートナーコードは既に使用されています"}), 400
                
                # 新規ユーザー作成
                new_user = User(
                    user_id=uuid.uuid4(),
                    partner_code=data['partner_code'],
                    password_hash=generate_password_hash(data['password']),
                    name=data['name'],
                    email=data.get('email'),
                    phone=data.get('phone'),
                    organization=data.get('organization'),
                    user_type=data['user_type'],
                    is_agency=data.get('is_agency', False),
                    is_active=data.get('is_active', True),
                    created_by=data.get('created_by', 'ADMIN'),
                    notes=data.get('notes')
                )
                
                db.session.add(new_user)
                db.session.commit()
                
                return jsonify({
                    "success": True,
                    "user": {
                        "user_id": str(new_user.user_id),
                        "partner_code": new_user.partner_code,
                        "name": new_user.name,
                    }
                }), 201
                
        except Exception as e:
            import traceback
            db.session.rollback()
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/admin/users/<user_id>", methods=['PUT', 'DELETE'])
    def admin_user_detail(user_id):
        """ユーザー管理API（PUT: 更新、DELETE: 削除）"""
        try:
            from models import User
            from sqlalchemy.exc import OperationalError
            from werkzeug.security import generate_password_hash
            import uuid
            
            # user_idをUUIDに変換
            try:
                user_id_uuid = uuid.UUID(user_id)
            except (ValueError, TypeError):
                return jsonify({"error": "無効なユーザーIDです"}), 400
            
            try:
                user = db.session.query(User).filter(User.user_id == user_id_uuid).first()
                
                if not user:
                    return jsonify({"error": "ユーザーが見つかりません"}), 404
                    
            except OperationalError as e:
                if 'no such table' in str(e).lower():
                    return jsonify({"error": "ユーザーテーブルが存在しません"}), 404
                raise
            
            if request.method == 'PUT':
                # ユーザー更新
                data = request.get_json()
                
                if not data:
                    return jsonify({"error": "リクエストボディが必要です"}), 400
                
                # パートナーコードの重複チェック（自分以外）
                if 'partner_code' in data and data['partner_code'] != user.partner_code:
                    existing = db.session.query(User).filter(
                        User.partner_code == data['partner_code'],
                        User.user_id != user_id_uuid
                    ).first()
                    
                    if existing:
                        return jsonify({"error": "このパートナーコードは既に使用されています"}), 400
                
                # フィールド更新
                if 'partner_code' in data:
                    user.partner_code = data['partner_code']
                if 'password' in data and data['password']:
                    user.password_hash = generate_password_hash(data['password'])
                if 'name' in data:
                    user.name = data['name']
                if 'email' in data:
                    user.email = data['email']
                if 'phone' in data:
                    user.phone = data['phone']
                if 'organization' in data:
                    user.organization = data['organization']
                if 'user_type' in data:
                    user.user_type = data['user_type']
                if 'is_agency' in data:
                    user.is_agency = data['is_agency']
                if 'is_active' in data:
                    user.is_active = data['is_active']
                if 'notes' in data:
                    user.notes = data['notes']
                
                db.session.commit()
                
                return jsonify({
                    "success": True,
                    "user": {
                        "user_id": str(user.user_id),
                        "partner_code": user.partner_code,
                        "name": user.name,
                    }
                })
                
            elif request.method == 'DELETE':
                # ユーザー削除（論理削除: is_activeをFalseに）
                user.is_active = False
                db.session.commit()
                
                return jsonify({"success": True, "message": "ユーザーを無効化しました"})
                
        except Exception as e:
            import traceback
            db.session.rollback()
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/export/excel")
    def export_excel():
        """ExcelエクスポートAPI"""
        try:
            from models import Store
            from flask import Response
            from sqlalchemy.exc import OperationalError
            
            try:
                query = _build_store_query()
                if query is None:
                    try:
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.append(['店舗ID', '店舗名'])
                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return Response(
                            output.getvalue(),
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': 'attachment; filename=stores_export.xlsx'}
                        )
                    except ImportError:
                        return jsonify({"error": "openpyxlライブラリがインストールされていません。"}), 500
                
                stores = query.all()
                
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, PatternFill
                    from openpyxl.utils import get_column_letter
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "店舗一覧"
                    
                    headers = [
                        '店舗ID', '店舗名', '電話番号', 'ウェブサイト', '住所', 'カテゴリ',
                        '評価', '都市', '開店日', '定休日', '交通アクセス', '営業時間',
                        '公式アカウント', 'データソース'
                    ]
                    ws.append(headers)
                    
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    for store in stores:
                        ws.append([
                            store.store_id, store.name, store.phone, store.website,
                            store.address, store.category, store.rating, store.city,
                            store.opening_date, store.closed_day, store.transport,
                            store.business_hours, store.official_account, store.data_source
                        ])
                    
                    for column in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    return Response(
                        output.getvalue(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'attachment; filename=stores_export.xlsx'}
                    )
                except ImportError:
                    return jsonify({"error": "openpyxlライブラリがインストールされていません。"}), 500
                    
            except OperationalError as e:
                if 'no such table' in str(e).lower():
                    try:
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.append(['店舗ID', '店舗名'])
                        output = io.BytesIO()
                        wb.save(output)
                        output.seek(0)
                        return Response(
                            output.getvalue(),
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': 'attachment; filename=stores_export.xlsx'}
                        )
                    except ImportError:
                        return jsonify({"error": "openpyxlライブラリがインストールされていません。"}), 500
                raise
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    def _build_store_query():
        """店舗クエリを構築（フィルターパラメータ対応）"""
        from models import Store
        from sqlalchemy import or_, and_
        from sqlalchemy.exc import OperationalError
        
        try:
            query = db.session.query(Store)
            
            # クエリパラメータ
            search = request.args.get("search", "").strip() or None
            search_mode = request.args.get("search_mode", "AND")
            match_type = request.args.get("match_type", "partial")
            prefectures = request.args.getlist("prefectures")
            cities = request.args.getlist("cities")
            categories = request.args.getlist("categories")
            data_sources = request.args.getlist("data_sources")
            
            # キーワード検索（店舗名・住所・カテゴリ）
            if search:
                keywords = [k for k in search.split() if k]
                if keywords:
                    conditions = []
                    for kw in keywords:
                        pattern = f"%{kw}%" if match_type == "partial" else kw
                        conditions.append(
                            or_(
                                Store.name.ilike(pattern),
                                Store.address.ilike(pattern),
                                Store.category.ilike(pattern),
                            )
                        )
                    if search_mode.upper() == "OR":
                        query = query.filter(or_(*conditions))
                    else:
                        for cond in conditions:
                            query = query.filter(cond)
            
            # 都道府県フィルター
            if prefectures:
                prefecture_filters = []
                for pref in prefectures:
                    prefecture_filters.append(
                        and_(
                            Store.address.isnot(None),
                            Store.address != "",
                            Store.address.like(f"{pref}%"),
                        )
                    )
                query = query.filter(or_(*prefecture_filters))
            
            # 市区町村フィルター
            if cities:
                query = query.filter(
                    Store.city.isnot(None),
                    Store.city != "",
                    Store.city.in_(cities),
                )
            
            # カテゴリフィルター（部分一致検索に対応）
            if categories:
                from sqlalchemy import or_
                # 選択されたカテゴリーがデータベースのカテゴリー値に含まれているかチェック
                # 例：「居酒屋」を選択した場合、「居酒屋」「松山市 / 居酒屋」などがマッチする
                category_filters = []
                for category in categories:
                    # 完全一致または部分一致（カテゴリー値に選択値が含まれる）
                    category_filters.append(
                        or_(
                            Store.category == category,
                            Store.category.contains(category)
                        )
                    )
                query = query.filter(
                    Store.category.isnot(None),
                    Store.category != "",
                    or_(*category_filters)
                )
            
            # データソースフィルター
            if data_sources:
                query = query.filter(
                    Store.data_source.isnot(None),
                    Store.data_source != "",
                    Store.data_source.in_(data_sources),
                )
            
            return query
        except OperationalError as e:
            if 'no such table' in str(e).lower():
                return None
            raise
    
    @app.route("/api/export/csv")
    def export_csv():
        """CSVエクスポートAPI"""
        try:
            from models import Store
            from flask import Response
            import csv
            import io
            from sqlalchemy.exc import OperationalError
            
            try:
                query = _build_store_query()
                if query is None:
                    # テーブルが存在しない場合は空のCSVを返す
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(['店舗ID', '店舗名'])
                    output.seek(0)
                    return Response(
                        output.getvalue(),
                        mimetype='text/csv; charset=utf-8',
                        headers={'Content-Disposition': 'attachment; filename=stores_export.csv'}
                    )
                
                stores = query.all()
                
                output = io.StringIO()
                writer = csv.writer(output)
                
                # ヘッダー
                writer.writerow([
                    '店舗ID', '店舗名', '電話番号', 'ウェブサイト', '住所', 'カテゴリ',
                    '評価', '都市', '開店日', '定休日', '交通アクセス', '営業時間',
                    '公式アカウント', 'データソース'
                ])
                
                # データ
                for store in stores:
                    writer.writerow([
                        store.store_id, store.name, store.phone, store.website,
                        store.address, store.category, store.rating, store.city,
                        store.opening_date, store.closed_day, store.transport,
                        store.business_hours, store.official_account, store.data_source
                    ])
                
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': 'attachment; filename=stores_export.csv'}
                )
            except OperationalError as e:
                if 'no such table' in str(e).lower():
                    # テーブルが存在しない場合は空のCSVを返す
                    output = io.StringIO()
                    writer = csv.writer(output)
                    writer.writerow(['店舗ID', '店舗名'])
                    output.seek(0)
                    return Response(
                        output.getvalue(),
                        mimetype='text/csv; charset=utf-8',
                        headers={'Content-Disposition': 'attachment; filename=stores_export.csv'}
                    )
                raise
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
    
    @app.route("/api/export/json")
    def export_json():
        """JSONエクスポートAPI"""
        try:
            from models import Store
            from flask import Response
            import json
            from sqlalchemy.exc import OperationalError
            
            try:
                query = _build_store_query()
                if query is None:
                    # テーブルが存在しない場合は空のJSONを返す
                    return Response(
                        json.dumps({"stores": []}, ensure_ascii=False, indent=2),
                        mimetype='application/json; charset=utf-8',
                        headers={'Content-Disposition': 'attachment; filename=stores_export.json'}
                    )
                
                stores = query.all()
                
                # 店舗データを辞書形式に変換
                stores_data = [store.to_dict() for store in stores]
                
                return Response(
                    json.dumps({"stores": stores_data}, ensure_ascii=False, indent=2),
                    mimetype='application/json; charset=utf-8',
                    headers={'Content-Disposition': 'attachment; filename=stores_export.json'}
                )
            except OperationalError as e:
                if 'no such table' in str(e).lower():
                    # テーブルが存在しない場合は空のJSONを返す
                    return Response(
                        json.dumps({"stores": []}, ensure_ascii=False, indent=2),
                        mimetype='application/json; charset=utf-8',
                        headers={'Content-Disposition': 'attachment; filename=stores_export.json'}
                    )
                raise
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
